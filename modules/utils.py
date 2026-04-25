import pandas as pd
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def detect_scale(series: pd.Series) -> str:
    try:
        s = series.dropna()
        if len(s) == 0:
            return "unknown"
        if not pd.api.types.is_numeric_dtype(series):
            return "categorical"
        u = s.unique()
        if len(u) <= 2:
            return "binary"
        if s.min() >= 1 and s.max() <= 7 and len(u) <= 7:
            return "likert"
        return "continuous"
    except Exception:
        return "unknown"


def cronbach_alpha(data: pd.DataFrame) -> float:
    try:
        d = data.select_dtypes(include=[np.number]).dropna()
        k = d.shape[1]
        if k < 2:
            return np.nan
        iv = d.var(ddof=1, axis=0).sum()
        tv = d.sum(axis=1).var(ddof=1)
        if tv == 0:
            return np.nan
        return round(float((k / (k - 1)) * (1 - iv / tv)), 3)
    except Exception:
        return np.nan


def calc_ave_cr(loadings):
    try:
        lam = np.array(loadings, dtype=float)
        lam2 = lam ** 2
        ave = float(np.mean(lam2))
        sl  = float(np.sum(lam))
        cr  = sl**2 / (sl**2 + float(np.sum(1 - lam2)))
        return round(ave, 3), round(cr, 3)
    except Exception:
        return np.nan, np.nan


def sig_stars(p) -> str:
    try:
        p = float(p)
        if p < 0.001: return "***"
        if p < 0.01:  return "**"
        if p < 0.05:  return "*"
        return "n.s."
    except Exception:
        return "-"


def fmt_p(p) -> str:
    try:
        p = float(p)
        return "< .001" if p < 0.001 else f"{p:.3f}"
    except Exception:
        return "-"


# ── Excel 스타일 ──────────────────────────────────────────────────────────────
def _hdr():
    return Font(name="Arial", bold=True, color="FFFFFF", size=11)

def _fill():
    return PatternFill("solid", fgColor="2F5496")

def _bd():
    return Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))

def _ctr():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _lft():
    return Alignment(horizontal="left", vertical="center")


def df_to_sheet(wb: Workbook, sheet_name: str, title: str,
                df_data: pd.DataFrame, note: str = "", adopt_col: str = None):
    try:
        ws = wb.create_sheet(sheet_name[:31])
        cols = list(df_data.columns)
        nc   = max(len(cols), 1)

        # 제목
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
        ws["A1"] = title
        ws["A1"].font = Font(name="Arial", bold=True, size=12)

        # 헤더
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=2, column=ci, value=str(col))
            c.font = _hdr(); c.fill = _fill()
            c.alignment = _ctr(); c.border = _bd()

        # adopt_col 인덱스
        adopt_idx = (cols.index(adopt_col) + 1) if adopt_col and adopt_col in cols else None

        # 데이터
        adopt_fill_y = PatternFill("solid", fgColor="E2EFDA")
        adopt_fill_n = PatternFill("solid", fgColor="FFC7CE")
        adopt_font_y = Font(name="Arial", size=10, color="375623")
        adopt_font_n = Font(name="Arial", size=10, color="9C0006")

        for ri, row in enumerate(df_data.itertuples(index=False), 3):
            for ci, val in enumerate(row, 1):
                safe_val = "" if (isinstance(val, float) and np.isnan(val)) else val
                cell = ws.cell(row=ri, column=ci, value=safe_val)
                cell.alignment = _lft() if ci <= 2 else _ctr()
                cell.border = _bd()
                if adopt_idx and ci == adopt_idx:
                    is_y = str(val) == "채택"
                    cell.fill = adopt_fill_y if is_y else adopt_fill_n
                    cell.font = adopt_font_y if is_y else adopt_font_n

        # 주석
        if note:
            nr = len(df_data) + 4
            note_cell = ws.cell(row=nr, column=1, value=note)
            note_cell.font = Font(name="Arial", size=9, color="595959", italic=True)

        # 열 너비 자동조정
        for col_cells in ws.columns:
            try:
                w = max((len(str(c.value)) if c.value else 0) for c in col_cells)
                ws.column_dimensions[col_cells[0].column_letter].width = min(w + 4, 45)
            except Exception:
                pass

        ws.row_dimensions[2].height = 28

    except Exception as e:
        pass  # 시트 생성 실패 시 조용히 넘어감


def build_excel(sheets: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for name, payload in sheets.items():
        try:
            title    = payload[0]
            df       = payload[1]
            note     = payload[2] if len(payload) > 2 else ""
            adopt_col = payload[3] if len(payload) > 3 else None
            if df is not None and len(df) > 0:
                df_to_sheet(wb, name, title, df, note, adopt_col)
        except Exception:
            pass
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
